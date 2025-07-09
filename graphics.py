import matplotlib.pyplot as plt
import os
import matplotlib.ticker as mticker
import openpyxl
from openpyxl import Workbook, load_workbook
from numpy import mean

def save_results(path,name, rewards, pollution=None, waits=None, bus=None):
    x = list(range(1, len(rewards) + 1))

    plot_dir  = os.path.join(path, "plots", name)



    # --------- Efficiency plot ---------
    plt.figure(figsize=(12, 5))
    plt.plot(x, rewards, marker='o', linestyle='-', markersize=3, color='blue')
    plt.title(f"{name}")
    plt.xlabel("Episode")
    plt.ylabel("Accumulated Reward")
    plt.grid(True)
    plt.tight_layout()
    plt.subplots_adjust(left=0.10)
    plt.gca().yaxis.set_major_formatter(mticker.ScalarFormatter())
    plt.gca().ticklabel_format(style='plain', axis='y')

    efficiency_path = f"{plot_dir}_rewards.png"
    plt.savefig(efficiency_path)
    plt.close()
    print(f"Efficiency plot saved to: {efficiency_path}")

    # --------- Pollution plot ---------
    type1=[]
    type2=[]
    type3=[]
    if pollution is not None:

        type1 = [p[0] for p in pollution]
        type2 = [p[1] for p in pollution]
        type3 = [p[2] for p in pollution]
        emissions=[p[0]+p[2] for p in pollution]
        plt.figure(figsize=(12, 5))
        plt.plot(x, type2, marker='s', linestyle='--', color='red', label='fuel')
        plt.plot(x, emissions, marker='^', linestyle='-.', color='green', label='Co2+NOx')
        #plt.plot(x, tipo3, marker='x', linestyle=':', color='green', label='Co2+NOx')

        plt.title(f"{name}")
        plt.xlabel("Episode")
        plt.ylabel("Pollution levels (mg)")
        plt.grid(True)
        plt.legend()
        plt.tight_layout()
        plt.subplots_adjust(left=0.10)
        plt.gca().yaxis.set_major_formatter(mticker.ScalarFormatter())
        plt.gca().ticklabel_format(style='plain', axis='y')

        pollution_path = f"{plot_dir}_pollution.png"
        plt.savefig(pollution_path)
        plt.close()
        print(f"Pollution plot saved to: {pollution_path}")

    # --------- Waiting time plot ---------
    if waits is not None:

        plt.figure(figsize=(12, 5))
        plt.plot(x, waits, marker='s', linestyle='-', color='orange')
        plt.title(f"{name}")
        plt.xlabel("Episode")
        plt.ylabel("Waiting time (s)")
        plt.grid(True)
        plt.tight_layout()
        plt.gca().yaxis.set_major_formatter(mticker.ScalarFormatter())
        plt.gca().ticklabel_format(style='plain', axis='y')

        wait_path = f"{plot_dir}_waiting_time.png"
        plt.savefig(wait_path)
        plt.close()
        print(f"Waiting time plot saved to: {wait_path}")

    # --------- Bus waiting time plot ---------
    if bus is not None:
        plt.figure(figsize=(12, 5))
        plt.plot(x, bus, marker='D', linestyle='-', color='purple')
        plt.title(f"{name}")
        plt.xlabel("Episode")
        plt.ylabel("Public transport waiting time (s)")
        plt.grid(True)
        plt.tight_layout()
        plt.gca().yaxis.set_major_formatter(mticker.ScalarFormatter())
        plt.gca().ticklabel_format(style='plain', axis='y')

        bus_path = f"{plot_dir}_bus_waiting_time.png"
        plt.savefig(bus_path)
        plt.close()
        print(f"Public transport plot saved to: {bus_path}")

    # --------- Guardar datos en Excel (.xlsx) ---------
    meanr = mean(rewards)
    minr = min(rewards)
    maxr = max(rewards)

    archivo_xlsx = os.path.join(path, "results.xlsx")
    hoja_nombre = "Results"

    needs_header = True
    if os.path.exists(archivo_xlsx):
        wb = load_workbook(archivo_xlsx)
        if hoja_nombre in wb.sheetnames:
            ws = wb[hoja_nombre]
            needs_header = False
        else:
            ws = wb.create_sheet(title=hoja_nombre)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = hoja_nombre

    #Headet
    expected_cols = 4 + len(rewards) * 6  # Name, Mean, Min, Max + 6 columns per episode
    if needs_header:
        ws.append([])
    if needs_header or ws.max_column != expected_cols:
        encabezado = ["Name", "Mean", "Minimum", "Maximum"]
        for i in range(len(rewards)):
            encabezado += [
                f"Reward_{i + 1}",
                f"CO2_{i + 1}",
                f"Fuel_{i + 1}",
                f"NOx_{i + 1}",
                f"Wait_{i + 1}",
                f"BusWait_{i + 1}",
            ]
        for col, value in enumerate(encabezado, start=1):
            ws.cell(row=1, column=col, value=value)

    # Fila de datos
    fila = [name, meanr, minr, maxr]
    for i in range(len(rewards)):
        fila.append(rewards[i])
        fila.append(type[i])
        fila.append(type[i])
        fila.append(type[i])
        fila.append(waits[i])
        fila.append(bus[i])
    ws.append(fila)

    wb.save(archivo_xlsx)
    print(f"Data saved to Excel file: {archivo_xlsx}")
